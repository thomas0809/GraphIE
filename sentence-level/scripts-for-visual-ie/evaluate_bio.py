import json
import sys
import os
import pandas as pd
import argparse
import numpy as np
import dateutil 
import datetime
import operator
from sacremoses import MosesTokenizer, MosesDetokenizer
from constants import *
import editdistance
import string
from tqdm import tqdm
from multiprocessing import Pool
from openpyxl import Workbook
from openpyxl.styles import Alignment
import copy

tok = MosesTokenizer()
detok = MosesDetokenizer()

parser = argparse.ArgumentParser(description='PyTorch')
parser.add_argument('--save_path', type=str, default='./')
parser.add_argument('--base', type=str, default='./')
parser.add_argument('--case', type=str, default='test_case.txt')
parser.add_argument('--model', type=str, default='lstm-gcn-lstm', choices=['lstm', 'lstm-lstm', 'lstm-gcn-lstm', 'lstm-rgcn-lstm', 'lstm-gat-lstm'])
parser.add_argument('--crf', action='store_true', default=False, help='final crf')
# parser.add_argument('--testmodel', type=str, default='')
parser.add_argument('--test', action='store_true', default=False, help='test mode')
args = parser.parse_args()

model_name = copy.copy(args.model)
if args.crf:
    model_name += '_crf'

if args.test:
    cmd = 'python train.py --test --save_path %s --case %s ' % (args.save_path, args.base+args.case)
    os.system(cmd)

excel = pd.read_excel('Cons_Full_training.xlsx')
record = {row.loc['Case Number']: row for idx, row in excel.iterrows()}

def process_single_word(candidate):
    count = {}
    for seq in candidate:   
        sent = detok.detokenize(seq, return_str=True)
        found = False
        for s in count:
            if s.lower() == sent.lower():
                count[s] += 1
                found = True
                break
        if not found:
            count[sent] = 1
    if len(count) == 0:
        return []
    _sorted = sorted(count.items(), key=operator.itemgetter(1), reverse=True)
    return [_sorted[0][0]]

def process_single_date(candidate):
    count = {}
    for seq in candidate:   
        sent = detok.detokenize(seq, return_str=True)
        if not isdate(sent):
            continue
        date = dateutil.parser.parse(sent)
        if date.year >= 2018:
            continue
        if date not in count:
            count[date] = 0
        count[date] += 1
    if len(count) == 0:
        return []
    _sorted = sorted(count.items(), key=operator.itemgetter(1), reverse=True)
    return [_sorted[0][0].strftime("%Y-%m-%d")]

def process_multip_sent(candidate):
    count = {}
    for seq in candidate:   
        sent = detok.detokenize(seq, return_str=True)
        found = False
        for s in count:
            s1 = ''.join([c for c in s.lower() if c not in string.punctuation])
            s2 = ''.join([c for c in sent.lower() if c not in string.punctuation])
            if s1 == s2:
                count[s] += 1
                found = True
                break
        if not found:
            count[sent] = 1
    return list(count.keys())

def process_multip_word(candidate):
    count = {}
    for seq in candidate:
        seq = [seq[0]]
        sent = detok.detokenize(seq, return_str=True)
        found = False
        for s in count:
            s1 = ''.join([c for c in s.lower() if c not in string.punctuation])
            s2 = ''.join([c for c in sent.lower() if c not in string.punctuation])
            if s1 == s2:
                count[s] += 1
                found = True
                break
        if not found:
            count[sent] = 1
    return list(count.keys())
    return process_multip_sent(candidate)


def calc_p_r_f1(x, y, ans, label):
    # p = float(sum(x))/len(ans) if len(ans)>0 else None
    # r = float(sum(y))/len(label) if len(label)>0 else None
    # return p, r
    # now return counts
    return sum(x), len(ans), sum(y), len(label)

def compare_word(ans, label, thres=1.0):
    ans = [w.lower() for w in ans]
    label = [w.lower() for w in label]
    x = [0 for w in ans]
    y = [0 for w in label]
    for i, w1 in enumerate(ans):
        for j, w2 in enumerate(label):
            if w1 == w2:
                x[i] = 1
                y[j] = 1
    return calc_p_r_f1(x, y, ans, label)

def compare_date(ans, label, thres=1.0):
    x = [0 for w in ans]
    y = [0 for w in label]
    for i, w1 in enumerate(ans):
        for j, w2 in enumerate(label):
            if samedate(w1, w2):
                x[i] = 1
                y[j] = 1
    return calc_p_r_f1(x, y, ans, label)

def similarity(s1, s2):
    seq1 = [w.lower() for w in tok.tokenize(s1) if w not in string.punctuation]
    seq2 = [w.lower() for w in tok.tokenize(s2) if w not in string.punctuation]
    s1 = detok.detokenize(seq1, return_str=True)
    s2 = detok.detokenize(seq2, return_str=True)
    ed = editdistance.eval(s1, s2)
    return 1.-float(ed)/max(len(s2),1)

def compare_sent(ans, label, thres=1.0):
    x = [0 for w in ans]
    y = [0 for w in label]
    for i, w1 in enumerate(ans):
        for j, w2 in enumerate(label):
            if similarity(w1, w2)>=thres:
                x[i] = 1
                y[j] = 1
    return calc_p_r_f1(x, y, ans, label)



def evaluate(path):
    try:
        if args.model in ['lstm', 'lstm-lstm']:
            filename = 'textline_bio.json'
        else:
            filename = 'graph_bio.json'
        graph = {}
        case_id = ''
        for line in open(path + filename):
            obj = json.loads(line)
            case_id = obj['id']['case']
            graph['%s %s %s'%(obj['id']['case'],obj['id']['doc'],obj['id']['page'])] = obj
        f = open(path + 'prob_%s.json'%model_name)
    except:
        return
    # attr_prec = {attr:{1.:[], 0.8:[], 0.6:[]} for attr in attrs}
    # attr_recall = {attr:{1.:[], 0.8:[], 0.6:[]} for attr in attrs}
    attr_stats = {attr:{1.:[0,0,0,0], 0.8:[0,0,0,0], 0.6:[0,0,0,0]} for attr in attrs}
    prediction = {}
    for line in f:
        try:
            obj = json.loads(line)
        except:
            print(path)
            return
        obj['pred'] = []
        for sent in obj['prob']:
            pred = []
            for word in sent:
                pred.append((np.argmax(word), np.max(word)))
            obj['pred'].append(pred)
        prediction['%s %s %s'%(obj['id']['case'],obj['id']['doc'],obj['id']['page'])] = obj
    candidates = {attr:[] for attr in attrs}
    for id, obj in prediction.items():
        for i, sent_pred in enumerate(obj['pred']):
            l = 0
            n = len(sent_pred)
            while l < n:
                tag_id, prob = sent_pred[l]
                if tag_id != 0:
                    attr = get_attr_from_tag_id(tag_id)
                    b_tag_id = tag2id[B_tag(attr)]
                    i_tag_id = tag2id[I_tag(attr)]
                    r = l
                    while r+1<n and sent_pred[r+1][0]==i_tag_id:
                        r += 1
                    span = graph[id]['sent'][i][l:r+1]
                    l = r
                    candidates[attr].append(span)
                l += 1
    output = {'id':case_id, 'ans':{}, 'candidate':{}}
    for attr in candidates:
        # attr = id2attr[id]
        anstype = attr2anstype[attr]
        if case_id in record:
            label = process_annotation(attr, record[case_id].loc[attr])
        else:
            label = []
        if anstype == SINGLE_WORD:
            ans = process_single_word(candidates[attr])
            compare_func = compare_word
        elif anstype == SINGLE_DATE:
            ans = process_single_date(candidates[attr])
            label_date = []
            for sent in label:
                try:
                    date = dateutil.parser.parse(sent)
                    label_date.append(date.strftime("%Y-%m-%d"))
                except:
                    continue
            label = label_date
            compare_func = compare_date
        elif anstype == MULTIP_WORD:
            ans = process_multip_word(candidates[attr])
            compare_func = compare_word
        elif anstype == MULTIP_SENT:
            ans = process_multip_sent(candidates[attr])
            compare_func = compare_sent
        for thres in [0.6, 0.8, 1.]:
            pred_hit, pred_sum, label_hit, label_sum = compare_func(ans, label, thres)
            x1, x2, x3, x4 = attr_stats[attr][thres]
            attr_stats[attr][thres] = [x1+pred_hit, x2+pred_sum, x3+label_hit, x4+label_sum]
            # if p is not None:
            #   attr_prec[attr][thres].append(p)
            # if r is not None:
            #   attr_recall[attr][thres].append(r)
        output['ans'][attr] = ans
        output['candidate'][attr] = candidates[attr]

    f = open(path + 'ans_%s.json'%model_name, 'w')
    json.dump(output, f)
    # print(path)
    return case_id, output['ans'], attr_stats #attr_prec, attr_recall
            
pool = Pool(processes=32)

f = open(args.base+args.case)
path_list = []
for idx, line in enumerate(f):
    path = args.base+line.strip()
    path_list.append(path)
    # evaluate(path, idx+2)

result = pool.map(evaluate, path_list)

# attr_prec = {attr:{1.:[], 0.8:[], 0.6:[]} for attr in attrs}
# attr_recall = {attr:{1.:[], 0.8:[], 0.6:[]} for attr in attrs}
attr_stats = {attr:{1.:[0,0,0,0], 0.8:[0,0,0,0], 0.6:[0,0,0,0]} for attr in attrs}

wb = Workbook()
ws = wb.active
ws['A1'] = 'Case Number'
for id, attr in id2attr.items():
    if id != 0:
        ws.cell(row=1, column=1+id).value = attr

row = 1
for res in result:
    if res is None:
        continue
    case_id, ans, stats = res
    row += 1
    ws.cell(row=row, column=1).value = case_id
    for id, attr in id2attr.items():
        if id != 0:
            ws.cell(row=row, column=1+id).value = ' \n'.join(ans[attr])
    for attr in attrs:
        for thres in [1.0, 0.8, 0.6]:
            pred_hit, pred_sum, label_hit, label_sum = stats[attr][thres]
            x1, x2, x3, x4 = attr_stats[attr][thres]
            attr_stats[attr][thres] = [x1+pred_hit, x2+pred_sum, x3+label_hit, x4+label_sum]
            # attr_prec[attr][thres] += prec[attr][thres]
            # attr_recall[attr][thres] += recall[attr][thres]

for col in ws.columns:
    column = col[0].column
    ws.column_dimensions[column].width = 25
wb.save(args.base+'predictions_%s.xlsx'%model_name)


ps, rs, f1s = [], [], []
for attr in attrs:
    if attr in ignore_attrs:
        continue
    for thres in [1.0]: #, 0.8, 0.6
        if attr2anstype[attr] not in [MULTIP_SENT] and thres != 1.:
            continue
        pred_hit, pred_sum, label_hit, label_sum = attr_stats[attr][thres]
        p = float(pred_hit)/pred_sum   #np.mean(attr_prec[attr][thres])
        r = float(label_hit)/label_sum #np.mean(attr_recall[attr][thres])
        f1 = 2*p*r/(p+r) #np.mean(attr_f1[attr][thres])
        print(attr, thres, 'prec=%.3f recall=%.3f f1=%.3f'%(p,r,f1))
        if thres == 1.:
            ps.append(p)
            rs.append(r)
            f1s.append(f1)
print(np.mean(ps), np.mean(rs), np.mean(f1s))

